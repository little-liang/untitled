import openpyxl
import MySQLdb
class mysql(object):
    """
    mysql 操作

    """
    def __init__(self):
        pass
    def common_run_sql(self, sql):
        config_info_json = {
        'host': "172.18.126.51",
        'port': 3306,
        'db': 'geetest',
        'user': 'root',
        'password': 'Abcd1234'}
        self.conn = MySQLdb.connect(**config_info_json, charset='utf8')
        self.cursor = self.conn.cursor()

        ## 待优化 每次查询都要连接断开,假设连接断开很慢,
        ## 最好本次程序运行完毕在进行数据库断开操作
        ##main 做成 上下文??
        try:
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            return result
        finally:
            self.conn.commit()
            self.cursor.close()
            self.conn.close()


    def __del__(self):
        pass

###实例一个表格对象
wb = openpyxl.load_workbook('hha.xlsx')

# print(wb.sheetnames)

##打开一个sheet
sheet = wb.get_sheet_by_name("Sheet1")

##第C列
# print(sheet["C"])

# print(sheet.max_row)
##第8行

run_sql_obj = mysql()

sql = "delete from geetest.query"
run_sql_obj.common_run_sql(sql)

for line in range(2, sheet.max_row + 1):
    tmp = "B%s" %(line)

    import datetime
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%Y%m%d%H%M%S")

    sql = "insert into geetest.query(name,flag,createtime,updatetime) value('%s','1','%s','%s')"% (sheet[tmp].value, now_time, now_time)

    run_sql_obj.common_run_sql(sql)

##G7的值
# print(sheet["G7"].value)